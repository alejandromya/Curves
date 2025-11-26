from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, PageBreak, Table, TableStyle
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import os

# ---------------------------
# Helpers para extraer campos
# ---------------------------
def _try_get(d: dict, keys, default=None):
    """Intentar obtener el primer key existente en d de la lista keys."""
    for k in keys:
        if k in d:
            return d[k]
    return default

def _format_num(v):
    if v is None:
        return "—"
    try:
        # Si es entero muy grande, mostrar sin decimales; si float, 2 decimales
        if isinstance(v, int):
            return f"{v}"
        return f"{float(v):.2f}"
    except Exception:
        return str(v)

# ---------------------------
# Generar una fila por sample
# ---------------------------
def generar_fila_sample(bloque):
    """
    Devuelve una lista con los valores en el orden de columnas:
    [Sample, 1,10,50,100,250,500, Cyclic Stiffness (N/mm),
     Yield Stiffness (N/mm), FMax ATM (N), Max Disp ATM (mm), Force at 2mm (N), Force at 3mm (N)]
    """
    df = bloque.get("df")
    detalles = bloque.get("ciclos", bloque.get("detalles", {})) or {}
    fuerza_max = bloque.get("fuerza_max")
    deformacion_max = bloque.get("deformacion_max")

    # ciclos objetivo (deformación HIGH)
    ciclos_obj = [1, 10, 50, 100, 250, 500]
    valores_ciclos = []

    for cnum in ciclos_obj:
        cd = detalles.get(cnum)
        if cd:
            # intento obtener deform_high_end o deform_high_start o deform_high
            deform_high = _try_get(cd, ["deform_high_end", "deform_high_start", "deform_high"])
            if deform_high is not None:
                valores_ciclos.append(_format_num(deform_high))
                continue
        valores_ciclos.append("—")

    # F2mm y F3mm: basados en LOW del último ciclo (intento obtener último ciclo)
    try:
        ciclos_ordenados = sorted(int(k) for k in detalles.keys())
    except Exception:
        # si keys no son enteras (improbable), fallback simple
        try:
            ciclos_ordenados = sorted([int(k) for k in detalles.keys() if str(k).isdigit()])
        except Exception:
            ciclos_ordenados = []

    if ciclos_ordenados:
        ultimo_ciclo = detalles[ciclos_ordenados[-1]]
        deform_low_last = _try_get(ultimo_ciclo, ["deform_low", "deform_low_start", "low_deform"])
    else:
        ultimo_ciclo = None
        deform_low_last = None

    f2mm_y = None
    f3mm_y = None
    if deform_low_last is not None and df is not None:
        try:
            f2mm_idx = (df["Deformacion"] - deform_low_last - 2.0).abs().idxmin()
            f3mm_idx = (df["Deformacion"] - deform_low_last - 3.0).abs().idxmin()
            f2mm_y = float(df.loc[f2mm_idx, "Fuerza"])
            f3mm_y = float(df.loc[f3mm_idx, "Fuerza"])
        except Exception:
            f2mm_y = None
            f3mm_y = None

    # Fmax relativo (fuerza máxima) y Max Disp ATM (deformacion_max)
    # Dciclico no se pide en tabla final; calculamos Yield (stf) según lo indicado: fuerza_max / (deformacion_max - deform_low_last)
    df_max_rel = None
    if deformacion_max is not None and deform_low_last is not None:
        try:
            df_max_rel = float(deformacion_max) - float(deform_low_last)
        except Exception:
            df_max_rel = None

    yield_stiffness = None
    if fuerza_max is not None and df_max_rel not in (None, 0):
        try:
            yield_stiffness = float(fuerza_max) / float(df_max_rel)
        except Exception:
            yield_stiffness = None

    # Cyclic stiffness: usar ciclo 250 si existe
    cyclic_stiffness = None
    cd250 = detalles.get(250)
    if cd250:
        # intentamos extraer forces y displacements desde cd250
        force_high = _try_get(cd250, ["force_high", "f_high", "fuerza_high", "force_peak", "peak_force"])
        force_low = _try_get(cd250, ["force_low", "f_low", "fuerza_low"])
        disp_high = _try_get(cd250, ["deform_high_end", "deform_high_start", "deform_high", "high_deform"])
        disp_low = _try_get(cd250, ["deform_low", "low_deform", "deform_low_start"])
        # si alguno falta, intentamos aproximar buscando en el DF: si dispones de índices guardados en cd250 (p. ej. high_idx, low_idx)
        if None in (force_high, force_low, disp_high, disp_low):
            # intentar leer campos alternativos que podrían guardar índices y buscar en df
            try:
                # intentos de índices
                high_idx = _try_get(cd250, ["high_idx", "deform_high_idx", "high_index"])
                low_idx = _try_get(cd250, ["low_idx", "deform_low_idx", "low_index"])
                if (high_idx is not None) and (low_idx is not None) and (df is not None):
                    # extraer desde df
                    try:
                        force_high = force_high or float(df.loc[int(high_idx), "Fuerza"])
                        force_low = force_low or float(df.loc[int(low_idx), "Fuerza"])
                        disp_high = disp_high or float(df.loc[int(high_idx), "Deformacion"])
                        disp_low = disp_low or float(df.loc[int(low_idx), "Deformacion"])
                    except Exception:
                        pass
            except Exception:
                pass

        # Si ya tenemos los cuatro valores, calcular
        if None not in (force_high, force_low, disp_high, disp_low):
            try:
                denom = float(disp_high) - float(disp_low)
                if denom != 0:
                    cyclic_stiffness = (float(force_high) - float(force_low)) / denom
            except Exception:
                cyclic_stiffness = None

    # Fmax ATM y Max Disp ATM
    fmax_atm = fuerza_max
    max_disp_atm = deformacion_max

    fila = []
    # Sample name
    fila.append(bloque.get("titulo", "Sample"))
    # columnas 1,10,50,100,250,500 (deform high)
    fila.extend(valores_ciclos)
    # Cyclic Stiffness
    fila.append(_format_num(cyclic_stiffness) if cyclic_stiffness is not None else "—")
    # Yield Stiffness (stf)
    fila.append(_format_num(yield_stiffness) if yield_stiffness is not None else "—")
    # FMax ATM
    fila.append(_format_num(fmax_atm) if fmax_atm is not None else "—")
    # Max Disp ATM
    fila.append(_format_num(max_disp_atm) if max_disp_atm is not None else "—")
    # Force at 2mm y 3mm
    fila.append(_format_num(f2mm_y) if f2mm_y is not None else "—")
    fila.append(_format_num(f3mm_y) if f3mm_y is not None else "—")

    return fila

# ---------------------------
# Excel: una hoja por columna
# ---------------------------
def agregar_hoja_excel(bloques, col_id, excel_path="INFORME_TOTAL.xlsx"):
    """
    bloques: lista de bloques de una columna (cada bloque = CSV)
    col_id: número de columna
    excel_path: ruta del Excel a crear/actualizar
    """
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        # eliminar hoja activa por defecto
        if wb.active:
            wb.remove(wb.active)

    ws = wb.create_sheet(title=f"Columna {col_id}")
    fila_actual = 1

    if not bloques:
        wb.save(excel_path)
        return excel_path

    # Cabecera fija
    headers = ["Sample", "1", "10", "50", "100", "250", "500",
               "Cyclic Stiffness (N/mm)", "Yield Stiffness (N/mm)",
               "FMax ATM (N)", "Max Disp ATM (mm)", "Force at 2mm (N)", "Force at 3mm (N)"]

    # Escribir cabecera
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=fila_actual, column=col_idx)
        cell.value = h
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    fila_actual += 1

    # Filas con datos
    for bloque in bloques:
        fila = generar_fila_sample(bloque)
        for col_idx, v in enumerate(fila, start=1):
            ws.cell(row=fila_actual, column=col_idx).value = v
        fila_actual += 1

    wb.save(excel_path)
    return excel_path

def generar_tablas_combinadas(bloques):
    """
    Genera dos tablas combinadas: Cyclic y 3rd Phase.
    Cada tabla tiene una fila por sample (bloque).
    """
    # Columnas Cyclic
    ciclos_col = [0, 10, 50, 100, 250, 500]
    headers_cyclic = [str(c) for c in ciclos_col] + ["Cyclic Stiffness (N/mm)"]

    # Columnas 3rd Phase
    headers_3rd = ["Yield Stiffness (N/mm)", "FMax ATM (N)", "Max Disp ATM (mm)", "Force at 2mm (N)", "Force at 3mm (N)"]

    filas_cyclic = []
    filas_3rd = []

    for bloque in bloques:
        df = bloque["df"]
        detalles_ciclos = bloque["ciclos"]
        fuerza_max = bloque["fuerza_max"]
        deformacion_max = bloque["deformacion_max"]

        primer_ciclo = detalles_ciclos[min(detalles_ciclos.keys())]
        ultimo_ciclo = detalles_ciclos[max(detalles_ciclos.keys())]

        # --- Cyclic ---
        fila_cyclic = []
        for c in ciclos_col:
            if c in detalles_ciclos:
                fila_cyclic.append(f"{detalles_ciclos[c]['deform_high_end']:.2f}")
            elif c == 0:
                fila_cyclic.append(f"{primer_ciclo['deform_high_start']:.2f}")
            else:
                fila_cyclic.append("—")
        df_max_rel = deformacion_max - ultimo_ciclo["deform_low"]
        cyclic_stiffness = fuerza_max / df_max_rel if df_max_rel != 0 else "—"
        fila_cyclic.append(f"{cyclic_stiffness:.2f}")
        filas_cyclic.append(fila_cyclic)

        # --- 3rd Phase ---
        f2mm_idx = (df["Deformacion"] - ultimo_ciclo["deform_low"] - 2.0).abs().idxmin()
        f3mm_idx = (df["Deformacion"] - ultimo_ciclo["deform_low"] - 3.0).abs().idxmin()
        f2mm_y = float(df.loc[f2mm_idx, "Fuerza"])
        f3mm_y = float(df.loc[f3mm_idx, "Fuerza"])
        yield_stiffness = ultimo_ciclo.get("stf", "—")
        max_disp_atm = deformacion_max

        fila_3rd = [
            f"{yield_stiffness:.2f}" if yield_stiffness != "—" else "—",
            f"{fuerza_max:.2f}",
            f"{max_disp_atm:.2f}",
            f"{f2mm_y:.2f}",
            f"{f3mm_y:.2f}"
        ]
        filas_3rd.append(fila_3rd)

    return (headers_cyclic, filas_cyclic), (headers_3rd, filas_3rd)


def generar_pdf_unico(bloques, output_pdf="INFORME_FINAL.pdf"):
    pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
    doc = SimpleDocTemplate(output_pdf, pagesize=letter)
    story = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='NormalUTF8', fontName='Arial', fontSize=10))

    story.append(Paragraph("<b>Resultados</b>", styles["Heading1"]))
    story.append(Spacer(1, 12))

    # Generar tablas combinadas
    (headers_cyclic, filas_cyclic), (headers_3rd, filas_3rd) = generar_tablas_combinadas(bloques)

    # Tabla Cyclic
    tabla_cyclic = Table([headers_cyclic] + filas_cyclic)
    tabla_cyclic.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
    ]))
    story.append(tabla_cyclic)
    story.append(Spacer(1, 12))

    # Tabla 3rd Phase
    tabla_3rd = Table([headers_3rd] + filas_3rd)
    tabla_3rd.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
    ]))
    story.append(tabla_3rd)
    story.append(PageBreak())

    # Ahora agregar gráficas por sample
    for bloque in bloques:
        story.append(Paragraph(f"<b>{bloque['titulo']}</b>", styles["Heading2"]))
        story.append(Spacer(1, 12))
        story.append(Image(bloque["grafico"], width=500, height=300))
        story.append(PageBreak())

    doc.build(story)
    return output_pdf
