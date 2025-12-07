from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from src.tables_generator import generar_fila_sample

def agregar_hoja_excel(bloques, col_id, excel_path_template="INFORME_COL{col}.xlsx"):
    """
    Crea un Excel por columna.
    Dentro del Excel, una hoja por sample.
    Cada hoja contiene la cabecera y una única fila de valores.
    """

    # Excel final para esta columna
    excel_path = excel_path_template.format(col=col_id)

    # Crear libro nuevo SIEMPRE (uno por columna)
    wb = Workbook()

    # Eliminar hoja inicial de openpyxl
    active_sheet = wb.active
    if active_sheet is not None:
        wb.remove(active_sheet)

    # Cabecera fija
    headers = [
        "Sample", "1", "10", "50", "100", "250", "500",
        "Cyclic Stiffness (N/mm)", "Yield Stiffness (N/mm)",
        "FMax ATM (N)", "Max Disp ATM (mm)",
        "Force at 2mm (N)", "Force at 3mm (N)"
    ]

    for bloque in bloques:
        # Nombre de hoja = Sample
        sample_name = bloque.get("titulo", "Sample")

        # Normalizar nombre (Excel no admite algunos caracteres)
        sample_name = sample_name.replace("/", "_").replace("\\", "_")

        ws = wb.create_sheet(title=sample_name)

        # Escribir cabecera
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = h
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Escribir fila de datos
        fila = generar_fila_sample(bloque)
        for col_idx, value in enumerate(fila, start=1):
            ws.cell(row=2, column=col_idx).value = value

    wb.save(excel_path)
    return excel_path
